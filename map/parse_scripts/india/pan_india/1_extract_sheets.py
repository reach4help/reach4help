# imports
import openpyxl
import os
import csv

# list files
path = "./_raw_data"
files = os.listdir(path)
print(files)

# Pick out 'xlsx' files
files_xlsx = [f for f in files if f[-4:] == "xlsx"]
print(files_xlsx)

# iterate over each xlsx file in the directory and then over worksheets
for file in files_xlsx:
    wb = openpyxl.load_workbook(os.path.join(path, file))
    print(wb.sheetnames)
    sheetnames = wb.sheetnames[1:]  # skipping the first sheet
    for sheetname in sheetnames:
        if "Sheet" not in sheetname:  # skipping empty sheets with "Sheet#"
            sheet = wb[sheetname]
            csvfilename = os.path.join(
                "input_files", file[:-5] + " - " + sheetname + ".csv"
            )
            col = csv.writer(open(csvfilename, "w", newline=""))
            for r in sheet.rows:
                # row by row write
                col.writerow([cell.value for cell in r])
